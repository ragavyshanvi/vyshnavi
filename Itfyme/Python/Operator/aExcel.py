import openpyxl
from openpyxl import Workbook
wb=openpyxl.load_workbook("C:\\ajeyaanna\\a1.xlsx")
wb2=openpyxl.load_workbook("C:\\ajeyaanna\\a2.xlsx")
sh1 = wb['Sheet1']
sh2 = wb2['Sheet1']
wb = openpyxl.Workbook()
wb['Sheet'].title= "name"
sh3 = wb.active
sh3['A1'].value =sh1['A1'].value+sh2['A1'].value
wb.save('h.xls')
print(sh3['A1'].value)
#data = sh1['A1'].value
#data2 =sh2['A1'].value
#print(type(sh1))
#print(type(wb))
#numSheets = wb.sheetnames
#print(numSheets)
#print(wb.active.title)
#print(data)
#print(data2)

